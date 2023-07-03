import tkinter as tk #Meng-import library U
import os #Meng-import library untuk mengecek file
import openpyxl as xl #Meng-import library untuk membaca file excel

#Inisiasi UI
windows = tk.Tk() #Inisiasi TKinter
windows.title("Aplikasi pencari data berdasarkan NIK") #Judul Program
windows.geometry("480x360") #Menentukan ukuran resolusi program
windows.resizable(False, False) #Mengunci resolusi program

#Deklarasi global variabel
sheet = None
text_error = None
text2 = None
text3 = None
text4 = None
text5 = None
text6 = None
text7 = None
text_loc = None
data_kategori = []
options = []
selected_option = tk.StringVar()
value_kat = None

def tiga_kategori(*args):
    global value_kat
    dropdown.configure(text="")
    value = selected_option.get()
    value = value.upper()
    if value == "KATEGORI KNN":
        value_kat = 4
    elif value == "KATEGORI NAIVE BAYES":
        value_kat = 5
    elif value == "KATEGORI DECISION TREE":
        value_kat = 6

#Function untuk mengecek apakah file berhasil dibuka
def pengecekan_file():
    global sheet, data_kategori
    varib =  [text_error, text2, text3, text4, text5, text_loc, text6, text7]
    input_text = namaFile.get()
    if input_text:
        if os.path.exists(input_text) and os.path.isfile(input_text):
            try:
                workbook = xl.load_workbook(input_text ,data_only=True)
                sheet = workbook.active
                workbook.close()
                tulisan = tk.Label(windows, text="File terbuka.\t\t\t", font=("Arial", 10))
                tulisan.place(x=75,y=38)
                tombol_klik1.config(state="normal") #Tombol kembali ke state normal
                nomorNik.config(state="normal")
                namaKategori.config(state='normal')
                dropdown.config(state='normal')
                data_kategori.clear()
                options.clear()
                data1 = []
                data1.clear()
                for i in varib:
                    if i is not None:
                        i.destroy()
                for row in sheet.iter_rows(min_row=1, max_row=1, min_col=5, max_col=7, values_only = True):
                    options.extend(row)
                dropdown['menu'].delete(0, 'end')
                for option in options:
                    dropdown['menu'].add_command(label=str(option), command=tk._setit(selected_option, option))
            except:
                pass
        else:
            tombol_klik1.config(state="disabled") #Tombol kembali ke state normal
            nomorNik.config(state="disabled")
            namaKategori.config(state='disabled')
            dropdown.config(state='disabled')
            tulisan = tk.Label(windows, text="Gagal untuk membuka file " + str(input_text) + " !!", font=("Arial", 10), foreground='red')
            tulisan.place(x=75,y=38)
    else:
        pass

#Function untuk mengecek apakah nik yang diinputkan user ada di file excel
def pengecekan_nik():
    global sheet, text_error, text2, text3, text4, text5, text_loc, text6, data_kategori, text7, value_kat
    varib =  [text_error, text2, text3, text4, text5, text_loc, text6, text7]
    nik = nomorNik.get() #Mengambil data yang diinput user dan menyimpannya di variabel nik
    var_kategori = namaKategori.get()
    var_kategori = var_kategori.upper()
    ditemukan =  False
    data_kategori.clear()
    for i in varib:
        if i is not None:
            i.destroy()
    #Melakukan perulangan untuk mendapatkan informasi tentang baris dan isi baris di file excel 
    for row_num, row in enumerate(sheet.iter_rows(min_row = 2, values_only = True), start=2):
        try:
            if str(row[0]) == str(nik): #Mengecek apakah nik yang diinput user ada di file
                text2 = tk.Label(windows, text="NIK           : " + str(row[0]))
                text2.place(x=25, y=168)
                text3 = tk.Label(windows, text="Berat badan : " + str(row[1]) + " Kg")
                text3.place(x=25, y=190)
                text4 = tk.Label(windows, text="Tinggi badan : " + str(row[2]) + " Cm")
                text4.place(x=25, y=212)
                text5 = tk.Label(windows, text="Jenis kelamin : " + str(row[3]))
                text5.place(x=25, y=234)
                text_loc = tk.Label(windows, text="Lokasi data berada dibaris ke " + str(row_num))
                text_loc.place(x=25, y=278)
                ditemukan = True
                text6 = tk.Label(windows, text="Kategori : " + str(row[value_kat]))
                text6.place(x=25, y=256)
            row_baru = []
            for i in row:
                if isinstance(i, str):
                    row_baru.append(i.upper())
                else:
                    row_baru.append(i)
            if str(row_baru[value_kat]) == var_kategori:
                data_kategori.append(row_baru)
        except:
            pass
    if nik:
        if ditemukan is not True :
            text_error = tk.Label(windows, text="Data dengan NIK " + str(nik) +" tidak ditemukan!!", foreground='red')
            text_error.place(x=125, y=338)
        else:
            text_error = tk.Label(windows, text="Data dengan NIK " + str(nik) +" ditemukan!!")
            text_error.place(x=125, y=338)
    else:
        pass
    if var_kategori:
        if len(data_kategori) > 0:
            text7 = tk.Label(windows, text="Terdapat "+ str(len(data_kategori)) + " data yang memiliki kategori " + str(var_kategori))
            text7.place(x=125, y=315)
        else:
            text7 = tk.Label(windows, text="Tidak ada data dengan kategori " + str(var_kategori), foreground='red')
            text7.place(x=125, y=315)
    else:
        pass

def hilangkan_text(event):
    dropdown.configure(text="")

#Global varioabel untuk menampilkan text, tombol, dan garis
garis = tk.Canvas(windows, width=455) #Membuat garis
garis.place(x=14, y=-88)
garis.create_line(0, 153, 447, 153)
garis.create_line(0, 230, 180, 230)
garis.create_line(250, 230, 447, 230)

label_pack = tk.Label(windows, text="@Halim648", font=("Arial", 9), relief='raised') #Membuat text untuk ditampilkan di apk
label_pack.place(x=406, y=338) #Meletakan text, tombol, dan kotak input sesuai dengan koordinat x dan y
label_pack.lift()

judul1 = tk.Label(windows, text="Masukan nama file : ", font=("Arial", 11))
judul1.place(x=25, y=15)

judul2 = tk.Label(windows, text="Masukan no NIK     : ", font=("Arial", 11))
judul2.place(x=25, y=105)

judul3 = tk.Label(windows, text="Info Data")
judul3.place(x=205, y=130)

judul4 = tk.Label(windows, text="Masukan Kategori  :", font=("Arial", 11))
judul4.place(x=25, y=73)

info1 = tk.Label(windows, text="Status : ", font=("Arial", 10))
info1.place(x=25,y= 38)

info2 = tk.Label(windows, text="Status kategori : ", font=("Arial", 10))
info2.place(x=25, y=315)

info3 = tk.Label(windows, text="Status NIK       : ", font=("Arial", 10))
info3.place(x=25, y=338)

namaFile = tk.Entry(windows, width=25) #Membuat sebuah kotak area untuk user menginput data (data dalam bentuk string)
namaFile.place(x=164, y=18)

nomorNik = tk.Entry(windows, width=25, state="disabled")
nomorNik.place(x=164, y=108)

namaKategori = tk.Entry(windows, width=25, state='disabled')
namaKategori.place(x=164, y=77)

tombol_klik = tk.Button(windows, text="Kirim", command=pengecekan_file) #Membuat tombol untuk mengirim data yang diinput user ke program
tombol_klik.place(x=330, y=13)

tombol_klik1 = tk.Button(windows, text="Kirim", command=pengecekan_nik, state="disabled") #Tombol dalam status tidak bisa ditekan
tombol_klik1.place(x=330, y=103)

area = tk.Label(windows, width=63, height=10, relief='sunken')
area.place(x=16, y=155)

dropdown = tk.OptionMenu(windows, selected_option, "")
dropdown.place(x=328, y=70)
dropdown.configure(state='disabled', width=5, text=" ")

selected_option.trace("w", tiga_kategori)

windows.mainloop() #Loop program